"""Extrai a aba Rumo-Pista da planilha para arquivos granulares de Escoteiro."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT.parent / "planilhaprogressao.xlsx"
TARGET = ROOT / "por" / "2025" / "escoteiro" / "planilha"


def slugify(text: str) -> str:
    """Gera um slug simples para nomes de arquivo."""
    replacements = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    text = text.lower()
    for old, new in replacements.items():
        text = text.replace(old, new)
    cleaned = []
    for ch in text:
        if ch.isalnum() or ch in {" ", "-"}:
            cleaned.append(ch)
        else:
            cleaned.append(" ")
    return "-".join(part for part in "".join(cleaned).split() if part)


def write_text(path: Path, text: str) -> None:
    """Grava um arquivo em UTF-8."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def normalize(value: object) -> str:
    """Converte celulas para texto util."""
    if value is None:
        return ""
    return str(value).strip()


def build_files() -> None:
    """Extrai a planilha em arquivos por bloco e por item."""
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Rumo-Pista"]

    grouped: dict[tuple[str, str, str], list[tuple[int, str]]] = defaultdict(list)
    rows = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        categoria = normalize(row[0])
        subcategoria = normalize(row[1])
        area = normalize(row[2])
        item = normalize(row[3])
        if categoria == "Categoria Macro":
            continue
        if not categoria or not subcategoria or not area or not item:
            continue
        grouped[(categoria, subcategoria, area)].append((idx, item))
        rows.append((idx, categoria, subcategoria, area, item))

    write_text(
        TARGET / "index.md",
        "\n".join(
            [
                "# Escoteiro - Rumo-Pista",
                "",
                "Fonte: `planilhaprogressao.xlsx` aba `Rumo-Pista`",
                f"Subtotal: {len(rows)} linhas de conteudo",
                "",
                "## Blocos",
                "",
            ]
        )
        + "\n",
    )

    index_lines = ["# Escoteiro - Rumo-Pista", "", "## Blocos", ""]
    for (categoria, subcategoria, area), items in sorted(grouped.items()):
        stage_slug = slugify(f"{categoria} {subcategoria} {area}")
        stage_path = TARGET / "etapas" / f"{stage_slug}.md"
        item_lines = [
            f"# {subcategoria}",
            "",
            f"Categoria Macro: `{categoria}`",
            f"Area: `{area}`",
            f"Fonte: `planilhaprogressao.xlsx`",
            f"Subtotal: {len(items)} itens",
            "",
            "## Itens",
            "",
        ]
        for row_number, item in items:
            item_lines.append(f"- Linha {row_number}: {item}")
        write_text(stage_path, "\n".join(item_lines).strip() + "\n")
        index_lines.append(
            f"- `{stage_slug}` - {categoria} / {subcategoria} / {area} - {len(items)} itens"
        )

        for row_number, item in items:
            item_slug = f"{row_number:03d}-{slugify(item.split(' - ', 1)[-1])}"
            item_path = TARGET / "itens" / f"{item_slug}.md"
            item_text = "\n".join(
                [
                    f"# {item}",
                    "",
                    f"Linha de origem: `{row_number}`",
                    f"Categoria Macro: `{categoria}`",
                    f"Subcategoria: `{subcategoria}`",
                    f"Area: `{area}`",
                    f"Fonte: `planilhaprogressao.xlsx`",
                    "",
                    "## Requisito",
                    "",
                    item,
                ]
            ) + "\n"
            write_text(item_path, item_text)

    write_text(TARGET / "index.md", "\n".join(index_lines).strip() + "\n")


if __name__ == "__main__":
    build_files()
