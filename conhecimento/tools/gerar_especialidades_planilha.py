"""Extrai a aba Especialidades da planilha para arquivos granulares."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT.parent / "planilhaprogressao.xlsx"
TARGET = ROOT / "especialidades" / "2025"


def slugify(text: str) -> str:
    """Gera um slug simples para nomes de arquivo."""
    replacements = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    text = text.lower()
    for old, new in replacements.items():
        text = text.replace(old, new)
    cleaned = []
    for ch in text:
        if ch.isalnum() or ch in {" ", "-"}:
            cleaned.append(ch)
        else:
            cleaned.append(" ")
    return "-".join(part for part in "".join(cleaned).split() if part)


def write_text(path: Path, text: str) -> None:
    """Grava um arquivo em UTF-8."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def normalize(value: object) -> str:
    """Converte celulas para texto util."""
    if value is None:
        return ""
    return str(value).strip()


def build_files() -> None:
    """Extrai a planilha em arquivos por ramo e por especialidade."""
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Especialidades"]

    grouped: dict[str, dict[str, list[tuple[int, str]]]] = defaultdict(lambda: defaultdict(list))
    total = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        ramo = normalize(row[0])
        especialidade = normalize(row[1])
        numero = normalize(row[2])
        requisito = normalize(row[3])
        if ramo == "Ramo de Conhecimento":
            continue
        if not ramo or not especialidade or not numero or not requisito:
            continue
        grouped[ramo][especialidade].append((idx, f"{numero} - {requisito}"))
        total += 1

    index_lines = [
        "# Especialidades 2025",
        "",
        "Fonte: `planilhaprogressao.xlsx` aba `Especialidades`",
        f"Subtotal: {total} linhas de conteudo",
        "",
        "## Ramos de conhecimento",
        "",
    ]

    for ramo, especialidades in sorted(grouped.items()):
        ramo_slug = slugify(ramo)
        ramo_dir = TARGET / "ramos" / ramo_slug
        ramo_lines = [
            f"# {ramo}",
            "",
            "Fonte: `planilhaprogressao.xlsx`",
            f"Subtotal: {sum(len(v) for v in especialidades.values())} itens",
            "",
            "## Especialidades",
            "",
        ]
        for especialidade, requisitos in sorted(especialidades.items()):
            esp_slug = slugify(especialidade)
            esp_path = ramo_dir / f"{esp_slug}.md"
            esp_lines = [
                f"# {especialidade}",
                "",
                f"Ramo de conhecimento: `{ramo}`",
                "Fonte: `planilhaprogressao.xlsx`",
                f"Subtotal: {len(requisitos)} requisitos",
                "",
                "## Requisitos",
                "",
            ]
            for row_number, requisito in requisitos:
                esp_lines.append(f"- Linha {row_number}: {requisito}")
            write_text(esp_path, "\n".join(esp_lines).strip() + "\n")
            ramo_lines.append(f"- `{esp_slug}` - {especialidade} - {len(requisitos)} requisitos")

        write_text(ramo_dir / "index.md", "\n".join(ramo_lines).strip() + "\n")
        index_lines.append(f"- `{ramo_slug}` - {ramo} - {sum(len(v) for v in especialidades.values())} requisitos")

    write_text(TARGET / "index.md", "\n".join(index_lines).strip() + "\n")


if __name__ == "__main__":
    build_files()
